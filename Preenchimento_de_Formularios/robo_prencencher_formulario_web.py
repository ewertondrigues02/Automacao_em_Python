# Importações
import time
from selenium import webdriver
import pandas as pd
from openpyxl import load_workbook 
from selenium.webdriver.common.by import By 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import pyautogui as mouse


# Inicialização do navegador
servico = Service()
options = webdriver.EdgeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Edge(service=servico, options=options)

# Site para prencher
driver.get('https://pt.surveymonkey.com/r/WLXYDX2')

# Diretórios dos Dados
arquivo = "dados_para_formulario.xlsx"
planilha_aberta = load_workbook(filename=arquivo)
celula_selecionada = planilha_aberta['dados']

# Para percorrer todos os dados da planilha usarei o laço "for"

while (celula_selecionada.max_row):
        for linha in range(2,len(celula_selecionada['A']) + 1):

            nome = celula_selecionada[f'A{linha}'].value
            email = celula_selecionada[f'B{linha}'].value
            sexo = celula_selecionada[f'C{linha}'].value
            telefone = celula_selecionada[f'D{linha}'].value
            sobreMim = celula_selecionada[f'E{linha}'].value


            espera = WebDriverWait(driver,10)

            # Tempo de espera
            time.sleep(3)

            
            # Prenche campo nome
            campo_nome = espera.until(ec.presence_of_element_located((By.XPATH,'//*[@id="166517069"]')))
            campo_nome.send_keys(nome)

            # Tempo de espera
            time.sleep(3)

            # Prenche o campo email
            campo_email = espera.until(ec.presence_of_element_located((By.XPATH,'//*[@id="166517072"]')))
            campo_email.send_keys(email)

            # Tempo de espera
            time.sleep(3)

            campo_telefone = espera.until(ec.presence_of_element_located((By.XPATH,'//*[@id="166517070"]')))
            campo_telefone.send_keys(telefone)

            # Tempo de espera
            time.sleep(3)

            # informar sexo
            campo_sexo = espera.until(ec.presence_of_element_located((By.CLASS_NAME,'radio-button-label-text')))
            campo_sexo.click()

            # Tempo de espera
            time.sleep(3)

            # Fale sobre você
            campo_sobreMim = espera.until(ec.presence_of_element_located((By.XPATH,'//*[@id="166517073"]')))
            campo_sobreMim.send_keys(sobreMim)

            # Tempo de espera
            time.sleep(3)

            if sexo == "Masculino":
                botao_mas = espera.until(ec.element_to_be_clickable((By.XPATH,'//*[@id="166517071_1215509812_label"]/span[2]')))
                botao_mas.click

            else:
                botao_fem = espera.until(ec.element_to_be_clickable((By.XPATH,'//*[@id="166517071_1215509813_label"]/span[2]')))
                botao_fem.click()
            

            # Confirma o botão de enviar
            botao_confirma = espera.until(ec.element_to_be_clickable((By.XPATH,'//*[@id="patas"]/main/article/section/form/div[2]/button')))
            botao_confirma.click()

            espera(driver,10)

