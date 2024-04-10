from openpyxl import load_workbook
import os

# Caminho do arquivo Excel
arquivo = "C:\\Users\Ewerton\OneDrive\\Documentos\Projetos\\Códigos do VS Code\\Robôs em Python\\Quebrar arquivo de Texto\\dados_dos_vendedores.xlsx"

# Carregar planilha
planilha_aberta = load_workbook(filename=arquivo)
celula_selecionada = planilha_aberta['dados']

# Variável para armazenar o nome anterior
nome_anterior = ""

# Obter o total de linhas na planilha
total_linhas = len(celula_selecionada['A']) + 1 

# Iterar sobre as linhas da planilha
for linha in range(2, total_linhas):
    # Obter o nome atual da célula na coluna A
    nome_atual = celula_selecionada[f'A{linha}'].value

    # Verificar se é o mesmo nome da iteração anterior
    if nome_anterior != nome_atual:
        # Criar nova planilha com o nome atual
        celula_resumo = planilha_aberta.create_sheet(title=nome_atual) 
        celula_selecionada_2 = planilha_aberta[nome_atual]

        # Adicionar cabeçalho à nova planilha
        celula_selecionada_2['A1'] = "Nome"
        celula_selecionada_2['B1'] = "E-mail"
        celula_selecionada_2['C1'] = "Profissão"
        celula_selecionada_2['D1'] = "Salário Hora"

        # Copiar dados para nova planilha
        celula_selecionada_2['A2'] = celula_selecionada[f'A{linha}'].value
        celula_selecionada_2['B2'] = celula_selecionada[f'B{linha}'].value
        celula_selecionada_2['C2'] = celula_selecionada[f'C{linha}'].value
        celula_selecionada_2['D2'] = celula_selecionada[f'D{linha}'].value

        # Atualizar nome anterior
        nome_anterior = nome_atual

# Salvar planilha
planilha_aberta.save(filename=arquivo)

# Abrir o arquivo Excel após a conclusão
os.startfile(arquivo)
